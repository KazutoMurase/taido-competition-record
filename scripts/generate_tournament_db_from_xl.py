#!/usr/bin/env python3
import click
import openpyxl
from openpyxl.cell import MergedCell
import os
import re


class Game:
    def __init__(self, id):
        self.id = id
        self.next_left = ''
        self.next_right = ''
        self.left_id = ''
        self.right_id = ''

    def __lt__(self, other):
        return self.id < other.id

    def __str__(self):
        return (f"{self.id},{self.left_id},{self.right_id},{self.next_left},{self.next_right},,,")


def refer_number(sheet, value):
    pattern = r"=([A-Z]+[0-9]+)\+([0-9]+)"
    match = re.match(pattern, str(value))
    if not match:
        return value
    ref_cell = match.group(1)
    addition = int(match.group(2))
    ref_value = sheet[ref_cell].value
    if isinstance(ref_value, int):
        return ref_value + addition
    return refer_number(sheet, ref_value) + addition


def check_number(sheet, row, col):
    if not isinstance(sheet[row][col], MergedCell):
        value = sheet[row][col].value
    elif (sheet[row][col-1].value is not None and
          (isinstance(sheet[row][col-1].value, int) or
           "=IF" not in sheet[row][col-1].value)):
        value = sheet[row][col-1].value
    elif (sheet[row-1][col].value is not None and
          (isinstance(sheet[row-1][col].value, int) or
           "=IF" not in sheet[row-1][col].value)):
        value = sheet[row-1][col].value
    elif (sheet[row][col-1].value is not None):
        return sheet[row][col-1].value
    else:
        return sheet[row-1][col].value
    number = refer_number(sheet, value)
    if isinstance(number, str) and '※' in number:
        number = int(number.replace('※', ''))
    return number


def check_towards_top(sheet, row, col):
    i = 1
    while (sheet[row-i][col-1].border.right.style == 'thin' or
           sheet[row-i][col].border.left.style == 'thin'):
        i += 1
    return i - 1


def check_towards_bottom(sheet, row, col):
    i = 1
    while (sheet[row+i][col-1].border.right.style == 'thin' or
           sheet[row+i][col].border.left.style == 'thin'):
        i += 1
    return i - 1


def check_towards_left(sheet, row, col):
    i = 1
    while (not isinstance(sheet[row+1][col-i], MergedCell) and
           (sheet[row][col-i].border.bottom.style == 'thin' or
           sheet[row+1][col-i].border.top.style == 'thin')):
        i += 1
    return i - 1


def check_towards_right(sheet, row, col):
    i = 1
    while (not isinstance(sheet[row+1][col+i], MergedCell) and
            (sheet[row][col+i].border.bottom.style == 'thin' or
             sheet[row+1][col+i].border.top.style == 'thin')):
        i += 1
    return i


def search_left_block(sheet, games, game, row, col):
    update_row = check_towards_top(sheet, row, col)
    update_col = check_towards_left(sheet, row - update_row-1, col)
    id = check_number(sheet, row - update_row - 1, col - update_col - 1)
    if isinstance(id, int):
        upper_game = Game(id=id)
        upper_game.next_left = game.id
        games.append(upper_game)
        search_left_block(sheet, games, upper_game,
                          row - update_row, col - update_col)
    elif isinstance(id, str):
        pattern = r"=IF\(\([A-Z]+[0-9]+"
        match = re.match(pattern, id)
        if match:
            ref_cell = match.group(0).replace('=IF((', '')
            game.left_id = sheet[ref_cell].value

    update_row = check_towards_bottom(sheet, row, col)
    update_col = check_towards_left(sheet, row + update_row, col)
    id = check_number(sheet, row + update_row, col - update_col - 1)
    if isinstance(id, int):
        lower_game = Game(id=id)
        lower_game.next_right = game.id
        games.append(lower_game)
        search_left_block(sheet, games, lower_game,
                          row + update_row, col - update_col)
    elif isinstance(id, str):
        pattern = r"=IF\(\([A-Z]+[0-9]+"
        match = re.match(pattern, id)
        if match:
            ref_cell = match.group(0).replace('=IF((', '')
            game.right_id = sheet[ref_cell].value


def search_right_block(sheet, games, game, row, col):
    update_row = check_towards_top(sheet, row, col)
    update_col = check_towards_right(sheet, row - update_row-1, col)
    id = check_number(sheet, row - update_row - 1, col + update_col)
    if isinstance(id, int):
        upper_game = Game(id=id)
        upper_game.next_right = game.id
        games.append(upper_game)
        search_right_block(sheet, games, upper_game,
                           row - update_row - 1, col + update_col)
    elif isinstance(id, str):
        pattern = r"=IF\(\([A-Z]+[0-9]+"
        match = re.match(pattern, id)
        if match:
            ref_cell = match.group(0).replace('=IF((', '')
            game.right_id = sheet[ref_cell].value

    update_row = check_towards_bottom(sheet, row, col)
    update_col = check_towards_right(sheet, row + update_row, col)
    id = check_number(sheet, row + update_row + 1, col + update_col)
    if isinstance(id, int):
        lower_game = Game(id=id)
        lower_game.next_left = game.id
        games.append(lower_game)
        search_right_block(sheet, games, lower_game,
                           row + update_row + 1, col + update_col)
    elif isinstance(id, str):
        pattern = r"=IF\(\([A-Z]+[0-9]+"
        match = re.match(pattern, id)
        if match:
            ref_cell = match.group(0).replace('=IF((', '')
            game.left_id = sheet[ref_cell].value

@click.command()
@click.option("--file-path", type=click.Path(exists=True, dir_okay=False), required=True)
@click.option("--output-path", type=click.Path(dir_okay=True), required=True)
def main(file_path, output_path):
    os.makedirs(output_path, exist_ok=True)
    wb = openpyxl.load_workbook(file_path)
    for name in wb.sheetnames:
        sheet = wb[name]
        start_cell = None
        for row in sheet.iter_rows():
            for cell in row:
                cell_value = cell.value if cell.value is not None else ''
                if cell_value == '決勝':
                    start_cell = cell
                    break
        if start_cell is None:
            continue
        games = []
        target_row, target_col = start_cell.row, start_cell.column
        index = check_towards_bottom(sheet, target_row, target_col)
        final_num = check_number(sheet, target_row+index+1, target_col)
        if final_num is None:
            continue
        # before_final / final
        before_final_game = Game(id=final_num-1)
        final_game = Game(id=final_num)
        games.append(before_final_game)
        games.append(final_game)
        left_index = check_towards_left(sheet, target_row+index, target_col)
        right_index = check_towards_right(sheet, target_row+index, target_col)
        # semi finals
        left_semi_final = Game(id=check_number(sheet, target_row+index, target_col-left_index-1))
        left_semi_final.next_left = final_num
        right_semi_final = Game(id=check_number(sheet, target_row+index, target_col+right_index))
        right_semi_final.next_right = final_num
        games.append(left_semi_final)
        games.append(right_semi_final)
        # search left/right block
        search_left_block(sheet, games, left_semi_final, target_row+index, target_col-left_index)
        search_right_block(sheet, games, right_semi_final, target_row+index, target_col+right_index)
        with open(f'{output_path}/{name}.csv', 'w') as f:
            if '団' in name:
                f.write('id,left_group_id,right_group_id,next_left_id,next_right_id,left_group_flag,left_retire,right_retire\n')
            else:
                f.write('id,left_player_id,right_player_id,next_left_id,next_right_id,left_player_flag,left_retire,right_retire\n')
            for game in sorted(games):
                f.write(f"{game}\n")


if __name__ == "__main__":
    main()
